import numpy as np
import os
from sklearn.metrics import matthews_corrcoef, f1_score, roc_curve
from sklearn import metrics
from sklearn.metrics import confusion_matrix, auc, precision_recall_curve
import openpyxl as op
import matplotlib.pyplot as plt
from model import get_model
import pandas as pd
from openpyxl import load_workbook
import warnings
warnings.filterwarnings("ignore")


Project_Path='insert your path'

def op_toexcel(data,filename):

    if os.path.exists(filename):
        wb = op.load_workbook(filename)
        ws = wb.worksheets[0]

        ws.append(data)
        wb.save(filename)
    else:
        wb = op.Workbook()
        ws = wb['Sheet']
        ws.append(['MCC', 'ACC', 'AUC', 'Sensitivity', 'Specificity', 'Precision', 'NPV', 'F1', 'FPR', 'FNR',
                  'TN', 'FP', 'FN', 'TP','AUPRC','Threshold'])
        ws.append(data)
        wb.save(filename)


def plot_roc_curve(y_true, y_pred):
    fpr, tpr, thresholds = metrics.roc_curve(y_true, y_pred)
    roc_auc = metrics.auc(fpr, tpr)

    plt.figure()
    lw = 2

    plt.plot(fpr, tpr, color='Red', lw=lw, label='ROC curve (area = %0.3f)' % roc_auc)
    plt.plot([0, 1], [0, 1], color='navy', lw=lw, linestyle='--')
    plt.xlim([0.0, 1.0])
    plt.ylim([0.0, 1.05])
    plt.xlabel('False Positive Rate', fontname='Arial')
    plt.ylabel('True Positive Rate', fontname='Arial')
    plt.tick_params(labelsize=10)
    plt.title('Receiver Operating Characteristic', fontsize=10)
    plt.legend(loc="lower right")
    plt.show()


def fcvtest(modelFile, filename, filename2, datasets):

    test_esm = np.load(Project_Path + f'features_npy/tentest/{datasets}/test_esmc_0.1.npy')
    test_pglm = np.load(Project_Path + f'features_npy/tentest/{datasets}/test_pglm_0.1.npy')
    test_eng = np.load(Project_Path + f'features_npy/tentest/{datasets}/test_eng_0.1.npy')
    test_label = np.load(Project_Path + f'features_npy/tentest/{datasets}/test_label_0.1.npy')

    train_model = get_model()
    train_model.load_weights(modelFile)

    y_pred = train_model.predict([test_pglm, test_esm, test_eng]).reshape(-1, )
    y_true =test_label
    y_pred_new = []

    best_f1 = 0
    best_threshold = 50
    for threshold in range(0, 1000):
        threshold = threshold / 1000
        binary_pred = [1 if pred >= threshold else 0 for pred in y_pred]
        f1 = metrics.f1_score(y_true, binary_pred)
        if f1 > best_f1:
            best_f1 = f1
            best_threshold = threshold
    for value in y_pred:
        if value < best_threshold:
            y_pred_new.append(0)
        else:
            y_pred_new.append(1)
    y_pred_new = np.array(y_pred_new)

    tn, fp, fn, tp = confusion_matrix(y_true, y_pred_new).ravel()
    fpr, tpr, thresholds = roc_curve(y_true, y_pred, pos_label=1)
    roc_auc = metrics.auc(fpr, tpr)
    precision, recall, thresholds = precision_recall_curve(y_true, y_pred_new)
    auprc = metrics.auc(recall, precision)
    thd =best_threshold

    print("Matthews: " + str(matthews_corrcoef(y_true, y_pred_new)))
    print("ACC: ", (tp + tn) / (tp + tn + fp + fn))
    print("AUC: ", roc_auc)
    print('sensitivity/recall:', tp / (tp + fn))
    print('specificity:', tn / (tn + fp))
    print('precision:', tp / (tp + fp))
    print('negative predictive value:', tn / (tn + fn))
    print("F1: " + str(f1_score(y_true, y_pred_new)))
    print('error rate:', fp / (tp + tn + fp + fn))
    print('false positive rate:', fp / (tn + fp))
    print('false negative rate:', fn / (tp + fn))
    print('TN:', tn, 'FP:', fp, 'FN:', fn, 'TP:', tp)
    print('AUPRC: ' + str(auprc))
    print('best_threshold: ' + str(best_threshold))

    mcc = float(format((matthews_corrcoef(y_true, y_pred_new)), '.4f'))
    acc = float(format((tp + tn) / (tp + tn + fp + fn), '.4f'))
    auc = float(format(roc_auc, '.4f'))
    sen = float(format(tp / (tp + fn), '.4f'))
    spe = float(format(tn / (tn + fp), '.4f'))
    pre = float(format(tp / (tp + fp), '.4f'))

    npv = float(format(tn / (tn + fn), '.4f'))
    f1 = float(format(f1_score(y_true, y_pred_new), '.4f'))
    fpr = float(format(fp / (tn + fp), '.4f'))
    fnr = float(format(fn / (tp + fn), '.4f'))
    auprc = float(format(auprc, '.4f'))

    result = mcc, acc, auc, sen, spe, pre, npv, f1, fpr, fnr, tn, fp, fn, tp, auprc, thd
    op_toexcel(result, filename)

    df = pd.DataFrame({
        'True': y_true,
        'Predict': y_pred,
        'Label':y_pred_new
    })

    try:
        book = load_workbook(filename2)
        book.remove(book.active)
    except FileNotFoundError:
        pass
    df.to_excel(filename2, index=False)